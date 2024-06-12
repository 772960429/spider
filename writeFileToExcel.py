#用openpyxl库，能处理的最大行数达到1048576,不支持xls格式,xlrd支持xls格式
#openpyxl 可以写入xls格式表中，无法新增数据
import openpyxl
from openpyxl.styles import Alignment
import os

class Openpyxl_Excel():
 
    def save_table(self, table, file_name):
        # 保存表
        table.save(file_name)
 
    def openpyxl_add(self, filepath=None, sheet_name=None, title=None,index=0):
        """
        :param filepath: excel路径
        :param sheet_name: 工作表名
        :param title: 标题
        :param index: 工作表的位置索引
        """
        if os.path.exists(filepath):
            workbook = openpyxl.load_workbook(filepath)
            sheet_names = workbook.sheetnames
            if sheet_name in sheet_names:
                wbsheet = workbook[sheet_name]
            else:
                wbsheet = workbook.create_sheet(title=sheet_name, index=0)
 
            max_row = wbsheet.max_row + 1
            if title:
                for i in range(0, len(title)):
                    wbsheet.cell(1, i + 1).value = title[i]
                    wbsheet.cell(1, i + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐
        else:
            workbook = openpyxl.Workbook()
            # 删除默认表Sheet， 创建指定表名
            if 'Sheet' in workbook.sheetnames and sheet_name != 'Sheet':
                del workbook["Sheet"]
                wbsheet = workbook.create_sheet(title=sheet_name, index=0)
            else:
                wbsheet = workbook[sheet_name]
            sheet_names = [sheet_name]
            if title:
                for i in range(0, len(title)):
                    wbsheet.cell(1, i+1).value = title[i]
                    wbsheet.cell(1, i + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐
            max_row = 2
        return workbook, sheet_names, wbsheet, max_row
 
    def add_data(self, file_name, title=None, datas=None, sheet_name=None):
        """
        :param file_name: 文件名
        :param title: sheet表标题 [title1, title2, title3]
        :param datas: 列表数据[[1,2,3],[4,5,6]]
        :param sheet_name: sheet表名
        """
        table, sheet_names, wbsheet, row = self.openpyxl_add(file_name, sheet_name, title)
        # 循环写入数据，居中对齐
        # datas = [[1, 2, 3], [1, 2, 3]]
        for i in range(len(datas)):
            for j in range(len(datas[i])):
                wbsheet.cell(row + i, j + 1).value = datas[i][j]  # 写入数据
                wbsheet.cell(row + i, j + 1).alignment = Alignment(horizontal='center', vertical='center')  # 居中对齐
        # 保存文件
        self.save_table(table, file_name)